import time
import requests
import json
from datetime import datetime
import openpyxl
import pyautogui
import pyperclip

pyautogui.FAILSAFE = True


def get_timestamp(timestr):  # 获取unix时间戳（毫秒）
    timestr = timestr[0:19]
    datetime_obj = datetime.strptime(timestr, "%Y-%m-%d %H:%M:%S")
    obj_stamp = int(time.mktime(datetime_obj.timetuple()) * 1000.0)
    return obj_stamp


def get_datetime(timeNum):  # 获取字符串时间戳（毫秒）
    timeStamp = float(timeNum / 1000)
    timeArray = time.localtime(timeStamp)
    timeStr = time.strftime("%Y-%m-%d %H:%M:%S", timeArray)
    return timeStr


def utc(local_datetime):
    local_unix = get_timestamp(local_datetime)
    utc_unix = local_unix - 28800000
    utc_datetime = str(get_datetime(utc_unix))
    return utc_datetime, utc_unix


def date_and_time(worksheet, r):
    try:
        date_time = worksheet.cell(row=r, column=1 + 1).value

        time = worksheet.cell(row=r, column=1 + 2).value
        hour = time.hour
        min = time.minute
        sec = time.second
        date_time = date_time.replace(hour=hour, minute=min, second=sec)

        return date_time

    except AttributeError:
        return worksheet.cell(row=r, column=1 + 1).value


def get_direction(worksheet, r):
    direction = worksheet.cell(row=r, column=1 + 5).value
    direction = direction.replace(' ', '')
    long_view = '看多'
    short_view = '看空'
    if direction == long_view:
        return True
    if direction == short_view:
        return False


def get_maxrow(worksheet):
    for i in range(1, worksheet.max_row + 2):
        while worksheet.cell(row=i + 1, column=1 + 5).value is None:
            return i
            break


def max_price(data3):
    list1 = []
    for i in range(len(data3)):
        list1.append(data3[i][2])
    price = max(list1)
    return price


def min_price(data3):
    list1 = []
    for i in range(len(data3)):
        list1.append(data3[i][3])
    price = min(list1)
    return price


def get_data(coin_type, local_start_unix, end_unix, interval):  # 请求一天内每三分钟的数据
    BASE_URL = 'https://api.binance.com'
    Kline = '/api/v1/klines'
    limit = 1000
    kline_url = BASE_URL + Kline + '?' + 'symbol=' + coin_type + '&interval=' + interval + '&startTime=' + str(
        local_start_unix) + '&endTime=' + str(end_unix) + '&limit=' + str(limit)
    resp = requests.get(kline_url)
    data = resp.json()
    return data


def signal_type(worksheet, r):
    if worksheet.cell(row=r, column=1 + 6).value is None and worksheet.cell(row=r,
                                                                            column=1 + 7).value is None:
        signalType = 'current'
    elif worksheet.cell(row=r, column=1 + 6).value is not None and worksheet.cell(row=r,
                                                                                  column=1 + 7).value is not None:
        signalType = 'interval'
    elif worksheet.cell(row=r, column=1 + 6).value is not None and worksheet.cell(row=r,
                                                                                  column=1 + 7).value is None:
        signalType = 'point'

    return signalType


def stop_loss_1(data, stop_line, direction):  # 是否触发止损

    if direction:

        touch_index = 0

        for i in range(len(data)):

            point_time = data[i][0]
            point_price = data[i][1]

            if float(point_price) <= stop_line:
                touch_status = True
                touch_time = point_time
                touch_index = i
                touch_price = point_price
                break
            else:
                touch_status = False
    else:
        touch_index = 0

        for i in range(len(data)):

            point_time = data[i][0]
            point_price = data[i][1]

            if float(point_price) >= stop_line:
                touch_status = True
                touch_time = point_time
                touch_index = i
                touch_price = point_price
                break
            else:
                touch_status = False

    return touch_status, touch_index


def stop_loss_2(data, stop_percent, start_price, direction):
    if direction:

        touch_index = 0

        for i in range(len(data)):

            point_time = data[i][0]
            point_price = data[i][1]
            point_percent = (float(point_price) - start_price) / start_price

            if point_percent <= (-1) * stop_percent:
                touch_status = True
                touch_index = i
                break
            else:
                touch_status = False
    else:

        touch_index = 0

        for i in range(len(data)):

            point_time = data[i][0]
            point_price = data[i][1]
            point_percent = (start_price - float(point_price)) / start_price

            if point_percent <= (-1) * stop_percent:
                touch_status = True
                touch_index = i
                break
            else:
                touch_status = False

    return touch_status, touch_index


def liquidation(data, start_price, leverage, direction):
    if direction:
        liquidation_index = 0

        for i in range(len(data)):
            point_unix = data[i][0]
            point_price = data[i][1]
            point_loss = (float(point_price) - start_price) / start_price
            point_loss_after_leverage = point_loss * leverage
            if point_loss_after_leverage <= -1:
                liquidation_status = True
                liquidation_index = i
                break
            else:
                liquidation_status = False
    else:
        liquidation_index = 0

        for i in range(len(data)):
            point_unix = data[i][0]
            point_price = data[i][1]
            point_loss = (start_price - float(point_price)) / start_price
            point_loss_after_leverage = point_loss * leverage
            if point_loss_after_leverage <= -1:
                liquidation_status = True
                liquidation_index = i
                break
            else:
                liquidation_status = False

    return liquidation_status, liquidation_index


def play_time_2side(data, lowerbound, upperbound):
    strike_index = 0

    for i in range(len(data)):

        point_price = data[i][1]
        point_index = i

        if (float(point_price) >= float(lowerbound)) and (float(point_price) <= float(upperbound)):
            strike_index = point_index
            play_or_not = True
            break
        else:
            play_or_not = False

    return play_or_not, strike_index


def play_time_1side(data, price_line, direction):
    strike_price = data[0][1]
    strike_index = 0

    if direction:

        for i in range(len(data)):

            point_time = data[i][0]
            point_price = data[i][1]
            point_index = i

            if float(point_price) < price_line:
                strike_price = point_price
                strike_index = point_index
                play_or_not = True
                break
            else:
                play_or_not = False
    else:

        for i in range(len(data)):

            point_price = data[i][1]
            point_index = i

            if float(point_price) > price_line:
                strike_price = point_price
                strike_index = point_index
                play_or_not = True
                break
            else:
                play_or_not = False

    return play_or_not, strike_index


def revise_start_position(signal, data_complete, lowerbound, upperbound):
    data_start_revised = data_complete
    play_status = False
    start_price = None

    signalType = signal['signalType']
    direction = signal['direction']
    if signalType == 'current':
        data_start_revised = data_complete
        start_time, start_price = get_start_time_and_price(data_start_revised)
        play_status = True

    elif signalType == 'point':
        play_status, play_index = play_time_1side(data_complete, lowerbound, direction)
        if play_status:
            data_start_revised = data_complete[play_index:]
            start_time, start_price = get_start_time_and_price(data_start_revised)
        else:
            signal['start_time'] = '未触发建议价格线'

    elif signalType == 'interval':
        play_status, play_index = play_time_2side(data_complete, lowerbound, upperbound)
        if play_status:
            data_start_revised = data_complete[play_index:]
            start_time, start_price = get_start_time_and_price(data_start_revised)
        else:
            signal['start_time'] = '未触发建议价格区间'

    return data_start_revised, start_price, play_status


def get_start_time_and_price(data_start_revised):
    start_price = float(data_start_revised[0][1])
    start_time = str(get_datetime(data_start_revised[0][0]))
    return start_time, start_price


def revise_end_position(signal, data_start_revised, start_price, stop_line, stop_percent, leverage):
    # 检测是否止损
    stop_loss_index = len(data_start_revised)
    stop_loss_time = None
    stop_loss_price = None

    direction = signal['direction']
    if stop_line is not None:
        stop_loss_status_1, stop_loss_index_1 = stop_loss_1(data_start_revised, stop_line, direction)
        if stop_loss_status_1:
            stop_loss_time = data_start_revised[stop_loss_index_1][0]
            stop_loss_price = data_start_revised[stop_loss_index_1][1]
            stop_loss_index = stop_loss_index_1
        else:
            stop_loss_index = len(data_start_revised)
            stop_loss_time = '未触及止损线'
            stop_loss_price = None
    elif stop_percent is not None:
        stop_loss_status_2, stop_loss_index_2 = stop_loss_2(data_start_revised, stop_percent, start_price, direction)
        if stop_loss_status_2:
            stop_loss_time = data_start_revised[stop_loss_index_2][0]
            stop_loss_price = data_start_revised[stop_loss_index_2][1]
            stop_loss_index = stop_loss_index_2
        else:
            stop_loss_index = len(data_start_revised)
            stop_loss_time = '未触及止损线'
            stop_loss_price = None
    else:
        stop_loss_index = len(data_start_revised)

    # 检测是否爆仓
    liquidation_status, liquidation_index_1 = liquidation(data_start_revised, start_price, leverage, direction)
    liquidation_index = len(data_start_revised)
    liquidation_time = None
    liquidation_price = None
    if liquidation_status:
        liquidation_time = data_start_revised[liquidation_index_1][0]
        liquidation_price = data_start_revised[liquidation_index_1][1]
        liquidation_index = liquidation_index_1
    else:
        liquidation_index = len(data_start_revised)
        liquidation_time = '未爆仓'
        liquidation_price = None

    # 判断哪个index小，哪个小用哪个切片数据
    end_revised_index = len(data_start_revised)

    if stop_loss_index < liquidation_index:
        end_revised_index = stop_loss_index

    elif stop_loss_index > liquidation_index:
        end_revised_index = liquidation_index

    elif stop_loss_index == liquidation_index:
        end_revised_index = stop_loss_index

    data_end_revised = data_start_revised[:end_revised_index + 1]

    return data_end_revised


def get_return(data, direction, start_price):
    if direction:
        max_ret = (float(data[-1][2]) - start_price) / start_price
    else:
        max_ret = (start_price - float(data[-1][2])) / start_price

    return max_ret


def get_a_signal(worksheet, r, x, y):
    signal = {}

    # 获取行号
    signal['row'] = r + 1000

    # 获取信号源
    signal['source'] = worksheet.cell(row=r, column=1).value

    # 获取開始时间
    local_start_datetime = str(date_and_time(worksheet, r))
    local_start_unix = get_timestamp(local_start_datetime)
    signal['local_start_datetime'] = local_start_datetime
    signal['local_start_unix'] = local_start_unix

    # 获取币种
    coin_type = worksheet.cell(row=r, column=1 + 3).value
    coin_type = coin_type.replace('/', '')
    signal['coin_pair'] = coin_type

    # 获取期限
    time_limit = worksheet.cell(row=r, column=5).value
    signal['time_limit'] = time_limit

    # 获取方向
    direction = get_direction(worksheet, r)
    signal['direction'] = direction
    signal['direction_str'] = worksheet.cell(row=r, column=6).value

    # 获取價格区间
    lowerbound = worksheet.cell(row=r, column=1 + 6).value
    upperbound = worksheet.cell(row=r, column=1 + 7).value
    signal['lowerbound'] = lowerbound
    signal['upperbound'] = upperbound

    # 获取价格中位数
    if worksheet.cell(row=r, column=1 + 6).value is not None and worksheet.cell(row=r, column=1 + 7).value is not None:
        signal['middle'] = (float(lowerbound) + float(upperbound)) / 2

    # 确定信号类型（确定入场价格）
    signalType = signal_type(worksheet, r)
    signal['signalType'] = signalType

    # 获取止损线
    stop_line = worksheet.cell(row=r, column=1 + 8).value
    stop_percent = worksheet.cell(row=r, column=1 + 9).value
    signal['stop_price'] = stop_line
    signal['stop_percentage'] = stop_percent

    # 获取杠杆
    leverage_str = worksheet.cell(row=r, column=1 + 10).value
    signal['leverage_str'] = leverage_str

    leverage = 1

    # 获取短期目标价
    short_term_aim = worksheet.cell(row=r, column=12).value
    signal['short_term_aim'] = short_term_aim

    not_found = load('not_found.json')
    if signal['row'] in not_found:
        signal['play_status'] = False
        return signal

    # 获取入场信息
    end_unix = local_start_unix + 259200000
    data = get_data(coin_type, local_start_unix, end_unix, interval='5m')
    data_start_revised, start_price, play_status = revise_start_position(signal, data, lowerbound, upperbound)
    signal['play_status'] = play_status

    if play_status:

        # 填写入场时间和价格
        signal['start_time'] = str(get_datetime(data_start_revised[0][0]))
        signal['start_price'] = str(start_price)
        expired = load('expired.json')
        start_alert = load('start_alert.json')
        if (signal['row'] not in start_alert)\
                and (signal['row'] not in expired) \
                and signal['signalType'] != 'current':
            start_reminder(signal, x, y)
            start_reminder_tg(signal)
            print('入场播报： ', signal['row'])
            add_data('start_alert.json', signal['row'])

    return signal


def get_a_signal_for_new_signal_dectect(worksheet, r):
    signal = {}

    # 获取行号
    signal['row'] = r + 1000

    # 获取開始时间
    local_start_datetime = str(date_and_time(worksheet, r))
    local_start_unix = get_timestamp(local_start_datetime)
    signal['local_start_datetime'] = local_start_datetime
    signal['local_start_unix'] = local_start_unix

    # 获取信号源
    signal['source'] = worksheet.cell(row=r, column=1).value

    # 获取币种
    coin_type = worksheet.cell(row=r, column=1 + 3).value
    coin_type = coin_type.replace('/', '')
    signal['coin_pair'] = coin_type

    # 获取期限
    time_limit = worksheet.cell(row=r, column=5).value
    signal['time_limit'] = time_limit

    # 获取方向
    signal['direction_str'] = worksheet.cell(row=r, column=6).value

    # 获取價格区间
    lowerbound = worksheet.cell(row=r, column=1 + 6).value
    upperbound = worksheet.cell(row=r, column=1 + 7).value
    signal['lowerbound'] = lowerbound
    signal['upperbound'] = upperbound

    # 获取止损线
    stop_line = worksheet.cell(row=r, column=1 + 8).value
    stop_percent = worksheet.cell(row=r, column=1 + 9).value
    signal['stop_price'] = stop_line
    signal['stop_percentage'] = stop_percent

    # 获取杠杆
    leverage_str = worksheet.cell(row=r, column=1 + 10).value
    signal['leverage_str'] = leverage_str

    # 获取短期目标价
    short_term_aim = worksheet.cell(row=r, column=12).value
    signal['short_term_aim'] = short_term_aim

    return signal


def update_a_signal(signal):
    local_start_unix = signal['local_start_unix']
    coin_type = signal['coin_pair']
    direction = signal['direction']
    signalType = signal['signalType']
    lowerbound = signal['lowerbound']
    upperbound = signal['upperbound']
    stop_line = signal['stop_price']
    stop_percent = signal['stop_percentage']

    play_status = signal['play_status']

    if play_status:
        start_price = float(signal['start_price'])
        current_local_unix = int(time.time()) * 1000  # 当前本地时间
        data = get_data(coin_type, current_local_unix - 3600000, current_local_unix, '1m')
        # 当前价格
        signal['current_price'] = float(data[-1][1])
        # 计算目前收益率
        updated_ret = get_return(data, direction, start_price)
        # 填入目前收益率
        signal['updated_return'] = updated_ret

    return signal


def get_signals_list(filename, x, y):
    workbook = openpyxl.load_workbook(filename)

    signals_list = []
    worksheet = workbook.active
    lower_row = get_maxrow(worksheet)
    upper_row = get_maxrow(worksheet) - 15
    for i in range(lower_row, max(1, upper_row), -1):
        signal = get_a_signal(worksheet, i, x, y)
        signals_list.append(signal)

    return signals_list, max(1, upper_row) + 1, lower_row


def request_signals_list(filename, x, y):
    initial_signals_list, upper_row, lower_row = get_signals_list(filename, x, y)
    signals_list = []
    for signal in initial_signals_list:
        signal = update_a_signal(signal)
        signals_list.append(signal)
    return signals_list, upper_row, lower_row


# save data to json file
def store(json_name, data):
    with open(json_name, 'w') as fw:
        json.dump(data, fw)


# load json data from file
def load(json_name):
    with open(json_name, 'r') as f:
        data = json.load(f)
        return data


# add new data into the json (read the old data and append the new data)
def add_data(jsonname, new_data):
    list1 = []
    with open(jsonname, 'r') as f:
        old_data = json.load(f)
        for item in old_data:
            list1.append(item)
    list1.append(new_data)
    data = list1
    with open(jsonname, 'w') as fw:
        json.dump(data, fw)


def refresh(filename, x, y):
    try:
        new_signals_list, new_upper_row, new_lower_row = request_signals_list(filename, x, y)

        # 检测是否超过预设盈利值
        row = load('rows.json')
        stop_alert = load('stop_price_alert.json')
        expired = load('expired.json')
        data_5 = load('five_percent.json')
        data_10 = load('ten_percent.json')
        data_20 = load('twenty_percent.json')
        data_30 = load('thirty_percent.json')
        data_40 = load('fourty_percent.json')
        data_50 = load('fifti_percent.json')
        data_60 = load('sixty_percent.json')
        data_70 = load('seventy_percent.json')
        data_80 = load('eighty_percent.json')
        data_90 = load('ninty_percent.json')
        data_100 = load('hundred_percent.json')
        data_middle = load('middle_alert.json')
        for i in range(len(new_signals_list)):
            play_status = new_signals_list[i]['play_status']
            if play_status:
                # 播报到达区间中点
                if new_signals_list[i]['signalType'] == 'interval' \
                        and (new_signals_list[i]['row'] not in expired) \
                        and (new_signals_list[i]['row'] not in data_middle) \
                        and (new_signals_list[i]['row'] not in data_5) \
                        and (
                        new_signals_list[i]['row'] >= (1000 + row - 6) and new_signals_list[i]['row'] <= (1000 + row)):
                    if new_signals_list[i]['direction'] and new_signals_list[i]['current_price'] <= new_signals_list[i][
                        'middle']:
                        middle_reminder(new_signals_list[i], x, y)
                        middle_reminder_tg(new_signals_list[i])
                        print('中点播报： ', new_signals_list[i]['row'])
                        add_data('middle_alert.json', new_signals_list[i]['row'])
                    if (not new_signals_list[i]['direction']) and new_signals_list[i]['current_price'] >= \
                            new_signals_list[i]['middle']:
                        middle_reminder(new_signals_list[i], x, y)
                        middle_reminder_tg(new_signals_list[i])
                        print('中点播报： ', new_signals_list[i]['row'])
                        add_data('middle_alert.json', new_signals_list[i]['row'])

                # 播报收益率
                if new_signals_list[i]['updated_return'] >= 0.05 \
                        and (new_signals_list[i]['row'] not in data_5) \
                        and (new_signals_list[i]['row'] not in stop_alert) \
                        and (new_signals_list[i]['row'] not in expired):
                    return_reminder(new_signals_list, i, '5%', x, y)
                    return_reminder(new_signals_list, i, '5%', 262, 367)
                    return_reminder_tg(new_signals_list, i, '5%')
                    print('5%播报： ', new_signals_list[i]['row'])
                    add_data('five_percent.json', new_signals_list[i]['row'])
                if new_signals_list[i]['updated_return'] >= 0.1 \
                        and (new_signals_list[i]['row'] not in data_10) \
                        and (new_signals_list[i]['row'] not in stop_alert) \
                        and (new_signals_list[i]['row'] not in expired):
                    return_reminder(new_signals_list, i, '10%', x, y)
                    return_reminder(new_signals_list, i, '10%', 262, 367)
                    return_reminder_tg(new_signals_list, i, '10%')
                    print('10%播报： ', new_signals_list[i]['row'])
                    add_data('ten_percent.json', new_signals_list[i]['row'])
                if new_signals_list[i]['updated_return'] >= 0.2 \
                        and (new_signals_list[i]['row'] not in data_20) \
                        and (new_signals_list[i]['row'] not in stop_alert) \
                        and (new_signals_list[i]['row'] not in expired):
                    return_reminder(new_signals_list, i, '20%', x, y)
                    return_reminder(new_signals_list, i, '20%', 262, 367)
                    return_reminder_tg(new_signals_list, i, '20%')
                    print('20%播报： ', new_signals_list[i]['row'])
                    add_data('twenty_percent.json', new_signals_list[i]['row'])
                if new_signals_list[i]['updated_return'] >= 0.3 \
                        and (new_signals_list[i]['row'] not in data_30) \
                        and (new_signals_list[i]['row'] not in stop_alert) \
                        and (new_signals_list[i]['row'] not in expired):
                    return_reminder(new_signals_list, i, '30%', x, y)
                    return_reminder(new_signals_list, i, '30%', 262, 367)
                    return_reminder_tg(new_signals_list, i, '30%')
                    print('30%播报： ', new_signals_list[i]['row'])
                    add_data('thirty_percent.json', new_signals_list[i]['row'])
                if new_signals_list[i]['updated_return'] >= 0.4 \
                        and (new_signals_list[i]['row'] not in data_40) \
                        and (new_signals_list[i]['row'] not in stop_alert) \
                        and (new_signals_list[i]['row'] not in expired):
                    return_reminder(new_signals_list, i, '40%', x, y)
                    return_reminder(new_signals_list, i, '40%', 262, 367)
                    return_reminder_tg(new_signals_list, i, '40%')
                    print('40%播报： ', new_signals_list[i]['row'])
                    add_data('fourty_percent.json', new_signals_list[i]['row'])
                if new_signals_list[i]['updated_return'] >= 0.5 \
                        and (new_signals_list[i]['row'] not in data_50) \
                        and (new_signals_list[i]['row'] not in stop_alert) \
                        and (new_signals_list[i]['row'] not in expired):
                    return_reminder(new_signals_list, i, '50%', x, y)
                    return_reminder(new_signals_list, i, '50%', 262, 367)
                    return_reminder_tg(new_signals_list, i, '50%')
                    print('50%播报： ', new_signals_list[i]['row'])
                    add_data('fifti_percent.json', new_signals_list[i]['row'])
                if new_signals_list[i]['updated_return'] >= 0.6 \
                        and (new_signals_list[i]['row'] not in data_60) \
                        and (new_signals_list[i]['row'] not in stop_alert) \
                        and (new_signals_list[i]['row'] not in expired):
                    return_reminder(new_signals_list, i, '60%', x, y)
                    return_reminder(new_signals_list, i, '60%', 262, 367)
                    return_reminder_tg(new_signals_list, i, '60%')
                    print('60%播报： ', new_signals_list[i]['row'])
                    add_data('sixty_percent.json', new_signals_list[i]['row'])
                if new_signals_list[i]['updated_return'] >= 0.7 \
                        and (new_signals_list[i]['row'] not in data_70) \
                        and (new_signals_list[i]['row'] not in stop_alert) \
                        and (new_signals_list[i]['row'] not in expired):
                    return_reminder(new_signals_list, i, '70%', x, y)
                    return_reminder(new_signals_list, i, '70%', 262, 367)
                    return_reminder_tg(new_signals_list, i, '70%')
                    print('70%播报： ', new_signals_list[i]['row'])
                    add_data('seventy_percent.json', new_signals_list[i]['row'])
                if new_signals_list[i]['updated_return'] >= 0.8 \
                        and (new_signals_list[i]['row'] not in data_80) \
                        and (new_signals_list[i]['row'] not in stop_alert) \
                        and (new_signals_list[i]['row'] not in expired):
                    return_reminder(new_signals_list, i, '80%', x, y)
                    return_reminder(new_signals_list, i, '80%', 262, 367)
                    return_reminder_tg(new_signals_list, i, '80%')
                    print('80%播报： ', new_signals_list[i]['row'])
                    add_data('eighty_percent.json', new_signals_list[i]['row'])
                if new_signals_list[i]['updated_return'] >= 0.9 \
                        and (new_signals_list[i]['row'] not in data_90) \
                        and (new_signals_list[i]['row'] not in stop_alert) \
                        and (new_signals_list[i]['row'] not in expired):
                    return_reminder(new_signals_list, i, '90%', x, y)
                    return_reminder(new_signals_list, i, '90%', 262, 367)
                    return_reminder_tg(new_signals_list, i, '90%')
                    print('90%播报： ', new_signals_list[i]['row'])
                    add_data('ninty_percent.json', new_signals_list[i]['row'])
                if new_signals_list[i]['updated_return'] >= 1 \
                        and (new_signals_list[i]['row'] not in data_100) \
                        and (new_signals_list[i]['row'] not in stop_alert) \
                        and (new_signals_list[i]['row'] not in expired):
                    return_reminder(new_signals_list, i, '100%', x, y)
                    return_reminder(new_signals_list, i, '100%', 262, 367)
                    return_reminder_tg(new_signals_list, i, '100%')
                    print('100%播报： ', new_signals_list[i]['row'])
                    add_data('hundred_percent.json', new_signals_list[i]['row'])



                # 播报止损
                if (new_signals_list[i]['row'] not in stop_alert) \
                    and (new_signals_list[i]['row'] not in data_5) \
                    and (new_signals_list[i]['row'] not in expired):
                    if new_signals_list[i]['direction'] \
                            and new_signals_list[i]['stop_price'] is not None \
                            and new_signals_list[i]['current_price'] <= float(new_signals_list[i]['stop_price']):
                        stop_price_alert(new_signals_list[i], x, y)
                        stop_price_alert_tg(new_signals_list[i])
                        print('止损%播报： ', new_signals_list[i]['row'])
                        add_data('stop_price_alert.json', new_signals_list[i]['row'])
                    if (not new_signals_list[i]['direction']) \
                            and new_signals_list[i]['stop_price'] is not None \
                            and new_signals_list[i]['current_price'] >= float(new_signals_list[i]['stop_price']):
                        stop_price_alert(new_signals_list[i], x, y)
                        stop_price_alert_tg(new_signals_list[i])
                        print('止损%播报： ', new_signals_list[i]['row'])
                        add_data('stop_price_alert.json', new_signals_list[i]['row'])

        monitor(new_upper_row, new_lower_row)
    except:
        alert(253, 201)


def monitor(new_upper_row, new_lower_row):
    pyautogui.hotkey('win', 'm')
    pyautogui.moveTo(253, 201)
    pyautogui.click()
    time.sleep(0.05)
    pyautogui.click()
    time.sleep(0.05)
    pyperclip.copy(
        "当前时间： " + str(get_datetime(int(time.time()) * 1000))[5:16] + '\n' +
        '当前追踪：【主流币】' + str(new_upper_row) + " - " + str(new_lower_row) + "行"
    )
    pyautogui.hotkey('ctrl', 'v')
    pyautogui.press('enter')
    time.sleep(0.1)
    pyautogui.hotkey('win', 'm')
    time.sleep(0.1)


def alert(x, y):
    pyautogui.hotkey('win', 'm')
    pyautogui.moveTo(x, y)
    pyautogui.click()
    time.sleep(0.05)
    pyautogui.click()
    time.sleep(0.05)
    pyperclip.copy(
        "当前时间： " + str(get_datetime(int(time.time()) * 1000))[5:16] + '\n' +
        '主流币信号追踪运行异常'
    )

    pyautogui.hotkey('ctrl', 'v')
    pyautogui.press('enter')
    time.sleep(0.1)
    pyautogui.hotkey('win', 'm')
    time.sleep(0.1)


def new_signal_remidner(signal, x, y):
    if signal['upperbound'] == None:
        signal['upperbound'] = ''
    if signal['lowerbound'] == None:
        signal['lowerbound'] = ''
    if signal['stop_price'] == None:
        signal['stop_price'] = ''
    if signal['stop_percentage'] == None:
        signal['stop_percentage'] = ''
    if signal['leverage_str'] == None:
        signal['leverage_str'] = ''
    if signal['short_term_aim'] == None:
        signal['short_term_aim'] = ''

    pyautogui.hotkey('win', 'm')
    pyautogui.moveTo(x, y)
    pyautogui.click()
    time.sleep(0.05)
    pyautogui.click()
    time.sleep(0.1)

    text1 = "【" + signal['coin_pair'].replace('USDT', '') + signal["time_limit"] + signal['direction_str'] + "】" + '\n'
    text9 = '信号编号： M' + str(signal['row']) + '\n'
    text2 = '交易对： ' + signal['coin_pair'] + ' ' + signal['leverage_str'] + '\n'
    text3 = '方向： ' + str(signal['direction_str']) + '\n'
    text4 = '建议入场价： ' + str(signal['lowerbound']) + ' - ' + str(signal['upperbound']) + '\n'
    text5 = '短期目标价： ' + str(signal['short_term_aim']) + '\n'
    text6 = '建议止损价： ' + str(signal['stop_price']) + str(signal['stop_percentage']) + '\n'
    text7 = '-------------' + '\n'
    text8 = '信号源： ' + signal['source']

    text = text1 + text9 + text2 + text3 + text4 + text5 + text6 + text7 + text8

    if signal['short_term_aim'] == '':
        text = text.replace(text5, '')
    if signal['stop_price'] == '' and signal['stop_percentage'] == '':
        text = text.replace(text6, '')
    if signal['lowerbound'] == '':
        text = text.replace(text4, '')

    pyperclip.copy(text)
    pyautogui.hotkey('ctrl', 'v')
    pyautogui.press('enter')
    time.sleep(0.1)
    pyautogui.hotkey('win', 'm')
    time.sleep(0.1)


def new_signal_remidner_tg(signal):
    if signal['upperbound'] == None:
        signal['upperbound'] = ''
    if signal['lowerbound'] == None:
        signal['lowerbound'] = ''
    if signal['stop_price'] == None:
        signal['stop_price'] = ''
    if signal['stop_percentage'] == None:
        signal['stop_percentage'] = ''
    if signal['leverage_str'] == None:
        signal['leverage_str'] = ''
    if signal['short_term_aim'] == None:
        signal['short_term_aim'] = ''

    pyautogui.hotkey('win', 'm')
    time.sleep(0.01)
    pyautogui.hotkey('win', '1')
    time.sleep(0.01)
    pyautogui.hotkey('ctrl', '1')
    time.sleep(0.01)

    text1 = "【" + signal['coin_pair'].replace('USDT', '') + signal["time_limit"] + signal['direction_str'] + "】" + '\n'
    text9 = '信號編號： M' + str(signal['row']) + '\n'
    text2 = '交易對： ' + signal['coin_pair'] + ' ' + signal['leverage_str'] + '\n'
    text3 = '方向： ' + str(signal['direction_str']) + '\n'
    text4 = '建議入場價： ' + str(signal['lowerbound']) + ' - ' + str(signal['upperbound']) + '\n'
    text5 = '短期目標價： ' + str(signal['short_term_aim']) + '\n'
    text6 = '建議止損價： ' + str(signal['stop_price']) + str(signal['stop_percentage']) + '\n'
    text7 = '-------------' + '\n'
    text8 = '信號源： ' + signal['source']

    text = text1 + text9 + text2 + text3 + text4 + text5 + text6 + text7 + text8

    if signal['short_term_aim'] == '':
        text = text.replace(text5, '')
    if signal['stop_price'] == '' and signal['stop_percentage'] == '':
        text = text.replace(text6, '')
    if signal['lowerbound'] == '':
        text = text.replace(text4, '')

    pyperclip.copy(text)
    pyautogui.hotkey('ctrl', 'v')
    pyautogui.press('enter')
    time.sleep(0.1)
    pyautogui.hotkey('win', 'm')
    time.sleep(0.1)


def start_reminder(signal, x, y):
    pyautogui.hotkey('win', 'm')
    pyautogui.moveTo(x, y)
    pyautogui.click()
    time.sleep(0.1)
    pyautogui.click()
    time.sleep(0.1)
    pyperclip.copy(
        "【" + signal['coin_pair'].replace('USDT', '') + "到达建议入场价】" + '\n' +
        '信号编号： M' + str(signal['row']) + ' (' + str(signal['local_start_datetime'])[5:10] + ')' + '\n' +
        '方向： ' + str(signal['direction_str']) + '\n' +
        '建议入场价： ' + str(signal['lowerbound']) + ' - ' + str(signal['upperbound']) + '\n' +
        '当前时间： ' + signal['start_time'][5:16] + '\n' +
        '当前入场价格： ' + str(signal['start_price']) + '\n'
    )
    pyautogui.hotkey('ctrl', 'v')
    pyautogui.press('enter')
    time.sleep(0.1)
    pyautogui.hotkey('win', 'm')
    time.sleep(0.1)


def start_reminder_tg(signal):
    pyautogui.hotkey('win', 'm')
    time.sleep(0.01)
    pyautogui.hotkey('win', '1')
    time.sleep(0.01)
    pyautogui.hotkey('ctrl', '1')
    time.sleep(0.01)

    pyperclip.copy(
        "【" + signal['coin_pair'].replace('USDT', '') + "到達建議入場價】" + '\n' +
        '信號編號： M' + str(signal['row']) + ' (' + str(signal['local_start_datetime'])[5:10] + ')' + '\n' +
        '方向： ' + str(signal['direction_str']) + '\n' +
        '建議入場價： ' + str(signal['lowerbound']) + ' - ' + str(signal['upperbound']) + '\n' +
        '當前時間： ' + signal['start_time'][5:16] + '\n' +
        '當前入場價格： ' + str(signal['start_price']) + '\n'
    )
    pyautogui.hotkey('ctrl', 'v')
    pyautogui.press('enter')
    time.sleep(0.1)
    pyautogui.hotkey('win', 'm')
    time.sleep(0.1)


def middle_reminder(signal, x, y):
    pyautogui.hotkey('win', 'm')
    pyautogui.moveTo(x, y)
    pyautogui.click()
    time.sleep(0.05)
    pyautogui.click()
    time.sleep(0.1)
    pyperclip.copy(
        "【" + signal['coin_pair'].replace('USDT', '') + "到达建议入场区间中点】" + '\n' +
        '信号编号： M' + str(signal['row']) + ' (' + str(signal['local_start_datetime'])[5:10] + ')' + '\n' +
        '方向： ' + str(signal['direction_str']) + '\n' +
        '建议入场价： ' + str(signal['lowerbound']) + ' - ' + str(signal['upperbound']) + '\n' +
        "当前时间： " + str(get_datetime(int(time.time()) * 1000))[5:16] + '\n' +
        '当前入场价格： ' + str(signal['current_price']) + '\n'
    )
    pyautogui.hotkey('ctrl', 'v')
    pyautogui.press('enter')
    time.sleep(0.1)
    pyautogui.hotkey('win', 'm')
    time.sleep(0.1)


def middle_reminder_tg(signal):
    pyautogui.hotkey('win', 'm')
    time.sleep(0.01)
    pyautogui.hotkey('win', '1')
    time.sleep(0.01)
    pyautogui.hotkey('ctrl', '1')
    time.sleep(0.01)

    pyperclip.copy(
        "【" + signal['coin_pair'].replace('USDT', '') + "到達建議入場區間中點】" + '\n' +
        '信號編號： M' + str(signal['row']) + ' (' + str(signal['local_start_datetime'])[5:10] + ')' + '\n' +
        '方向： ' + str(signal['direction_str']) + '\n' +
        '建議入場價： ' + str(signal['lowerbound']) + ' - ' + str(signal['upperbound']) + '\n' +
        "當前時間： " + str(get_datetime(int(time.time()) * 1000))[5:16] + '\n' +
        '當前入場價格： ' + str(signal['current_price']) + '\n'
    )
    pyautogui.hotkey('ctrl', 'v')
    pyautogui.press('enter')
    time.sleep(0.1)
    pyautogui.hotkey('win', 'm')
    time.sleep(0.1)


def stop_price_alert(signal, x, y):
    pyautogui.hotkey('win', 'm')
    pyautogui.moveTo(x, y)
    pyautogui.click()
    time.sleep(0.05)
    pyautogui.click()
    time.sleep(0.1)
    pyperclip.copy(
        "【" + signal['coin_pair'].replace('USDT', '') + "触发止损】" + '\n' +
        '信号编号： M' + str(signal['row']) + ' (' + str(signal['local_start_datetime'])[5:10] + ')' + '\n' +
        '方向： ' + str(signal['direction_str']) + '\n' +
        '建议止损价： ' + str(signal['stop_price']) + '\n' +
        "当前时间： " + str(get_datetime(int(time.time()) * 1000))[5:16] + '\n' +
        '当前价格： ' + str(signal['current_price']) + '\n'
    )
    pyautogui.hotkey('ctrl', 'v')
    pyautogui.press('enter')
    time.sleep(0.1)
    pyautogui.hotkey('win', 'm')
    time.sleep(0.1)


def stop_price_alert_tg(signal):
    pyautogui.hotkey('win', 'm')
    time.sleep(0.01)
    pyautogui.hotkey('win', '1')
    time.sleep(0.01)
    pyautogui.hotkey('ctrl', '1')
    time.sleep(0.01)

    pyperclip.copy(
        "【" + signal['coin_pair'].replace('USDT', '') + "觸發止損】" + '\n' +
        '信號編號： M' + str(signal['row']) + ' (' + str(signal['local_start_datetime'])[5:10] + ')' + '\n' +
        '方向： ' + str(signal['direction_str']) + '\n' +
        '建議止損價： ' + str(signal['stop_price']) + '\n' +
        "當前時間： " + str(get_datetime(int(time.time()) * 1000))[5:16] + '\n' +
        '當前價格： ' + str(signal['current_price']) + '\n'
    )
    pyautogui.hotkey('ctrl', 'v')
    pyautogui.press('enter')
    time.sleep(0.1)
    pyautogui.hotkey('win', 'm')
    time.sleep(0.1)


def return_reminder(new_signals_list, i, ret, x, y):
    pyautogui.hotkey('win', 'm')
    pyautogui.moveTo(x, y)
    pyautogui.click()
    time.sleep(0.05)
    pyautogui.click()
    time.sleep(0.1)
    pyperclip.copy(
        "【" + new_signals_list[i]['coin_pair'].replace('USDT', '') + '收益突破' + ret + "】" + '\n' +
        '信号编号： M' + str(new_signals_list[i]['row']) + ' (' + str(new_signals_list[i]['local_start_datetime'])[
                                                             5:10] + ')' + '\n' +
        '推荐时间： ' + new_signals_list[i]['local_start_datetime'][5:16] + '\n' +
        "当前时间： " + str(get_datetime(int(time.time()) * 1000))[5:16] + '\n' +
        "当前价格： " + str(new_signals_list[i]['current_price']) + '\n' +
        '当前收益： {:.2%}'.format(new_signals_list[i]['updated_return']) + '\n'
    )
    pyautogui.hotkey('ctrl', 'v')
    pyautogui.press('enter')
    time.sleep(0.1)
    pyautogui.hotkey('win', 'm')
    time.sleep(0.1)


def return_reminder_tg(new_signals_list, i, ret):
    pyautogui.hotkey('win', 'm')
    time.sleep(0.01)
    pyautogui.hotkey('win', '1')
    time.sleep(0.01)
    pyautogui.hotkey('ctrl', '1')
    time.sleep(0.01)

    pyperclip.copy(
        "【" + new_signals_list[i]['coin_pair'].replace('USDT', '') + '收益突破' + ret + "】" + '\n' +
        '信號編號： M' + str(new_signals_list[i]['row']) + ' (' + str(new_signals_list[i]['local_start_datetime'])[
                                                             5:10] + ')' + '\n' +
        '推薦時間： ' + new_signals_list[i]['local_start_datetime'][5:16] + '\n' +
        "當前時間： " + str(get_datetime(int(time.time()) * 1000))[5:16] + '\n' +
        "當前價格： " + str(new_signals_list[i]['current_price']) + '\n' +
        '當前收益： {:.2%}'.format(new_signals_list[i]['updated_return']) + '\n'
    )
    pyautogui.hotkey('ctrl', 'v')
    pyautogui.press('enter')
    time.sleep(0.1)
    pyautogui.hotkey('win', 'm')
    time.sleep(0.1)


def expired_reminder(signal, x, y):
    pyautogui.hotkey('win', 'm')
    pyautogui.moveTo(x, y)
    pyautogui.click()
    time.sleep(0.05)
    pyautogui.click()
    time.sleep(0.1)
    pyperclip.copy(
        '【信号M' + str(signal['row']) + ' (' + str(signal['local_start_datetime'])[5:10] + ')' + '过期】' + '\n' +
        '该币' + signal['coin_pair'].replace('USDT', '') + '已有新信号发出，请参考最新信号的建议'
    )
    pyautogui.hotkey('ctrl', 'v')
    pyautogui.press('enter')
    time.sleep(0.1)
    pyautogui.hotkey('win', 'm')
    time.sleep(0.1)


def expired_reminder_tg(signal):
    pyautogui.hotkey('win', 'm')
    time.sleep(0.01)
    pyautogui.hotkey('win', '1')
    time.sleep(0.01)
    pyautogui.hotkey('ctrl', '1')
    time.sleep(0.01)

    pyperclip.copy(
        '【信號M' + str(signal['row']) + ' (' + str(signal['local_start_datetime'])[5:10] + ')' + '過期】' + '\n' +
        '該幣' + signal['coin_pair'].replace('USDT', '') + '已有新信號發出，請參考最新信號的建議'
    )
    pyautogui.hotkey('ctrl', 'v')
    pyautogui.press('enter')
    time.sleep(0.1)
    pyautogui.hotkey('win', 'm')
    time.sleep(0.1)


def new_signal_dectect(filename, x, y):
    signal_list_for_new_signal_dectect = []
    workbook = openpyxl.load_workbook(filename)
    worksheet = workbook.active
    new_lower_row = get_maxrow(worksheet)
    old_lower_row = load('rows.json')
    old_upper_row = old_lower_row - 15
    new_upper_row = get_maxrow(worksheet) - 15
    current_signal_list = []
    expired = load('expired.json')
    for r in range(old_upper_row, old_lower_row + 1):
        s = get_a_signal_for_new_signal_dectect(worksheet, r)
        current_signal_list.append(s)

    if new_lower_row != old_lower_row:
        diff = new_lower_row - old_lower_row
        for i in range(new_lower_row, new_lower_row - diff, -1):
            signal = get_a_signal_for_new_signal_dectect(worksheet, i)
            signal_list_for_new_signal_dectect.append(signal)
        for i in range(0, diff):
            new_signal_remidner(signal_list_for_new_signal_dectect[i], x, y)
            new_signal_remidner_tg(signal_list_for_new_signal_dectect[i])
            for s in current_signal_list:
                if (s['row'] not in expired) and s['coin_pair'] == signal_list_for_new_signal_dectect[i]['coin_pair']:
                    expired_reminder(s, x, y)
                    expired_reminder_tg(s)
                    add_data('expired,json', s['row'])

    store('rows.json', new_lower_row)

# 全部杠杆为1，没有爆仓可能。