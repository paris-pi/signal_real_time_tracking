import openpyxl
from datetime import datetime
import time
import requests
import json
import pyautogui
import pyperclip
from threading import Timer


def get_datetime(timeNum):  # 获取字符串时间戳（毫秒）
    timeStamp = float(timeNum / 1000)
    timeArray = time.localtime(timeStamp)
    timeStr = time.strftime("%Y-%m-%d %H:%M:%S", timeArray)
    return timeStr

def get_utc_timestamp(timestr):  # 获取unix时间戳（毫秒）
    timestr = timestr[0:19]
    datetime_obj = datetime.strptime(timestr, "%Y-%m-%d %H:%M:%S")
    obj_stamp = int(time.mktime(datetime_obj.timetuple()))
    return obj_stamp - 28800


def get_maxrow(worksheet):
    for i in range(1, worksheet.max_row + 2):
        while worksheet.cell(row=i + 1, column=1).value is None:
            return i
            break


# save data to json file
def store(json_name,data):
    with open(json_name, 'w') as fw:
        json.dump(data,fw)
# load json data from file
def load(json_name):
    with open(json_name,'r') as f:
        data = json.load(f)
        return data
# add new data into the json (read the old data and append the new data)
def add_data(jsonname,new_data):
    list1 = []
    with open(jsonname,'r') as f:
        old_data = json.load(f)
        for item in old_data:
            list1.append(item)
    list1.append(new_data)
    data = list1
    with open(jsonname, 'w') as fw:
        json.dump(data,fw)



def get_signal(worksheet,r):

    signal = {}
    signal['row'] = r +1000
    signal['datetime'] = worksheet.cell(row=r,column=1).value
    signal['coin_pair'] =  worksheet.cell(row=r,column=2).value
    signal['network'] =  worksheet.cell(row=r,column=3).value
    signal['token_id'] =  worksheet.cell(row=r,column=4).value
    signal['pair_id'] = worksheet.cell(row=r,column=5).value
    signal['utc_now'] = datetime.utcnow().isoformat()
    utc_start = datetime.fromtimestamp(get_utc_timestamp(str(worksheet.cell(row=r,column=1).value))).isoformat()
    signal['utc_start'] = utc_start
    query = get_query_first(signal)

    data = run_query(query)['data']['ethereum']["dexTrades"][0]
    signal['start_price'] = data['quotePrice']
    signal['start_time'] = data['timeInterval']['second']
    signal['base_currency'] = data['baseCurrency']['symbol']
    signal['quote_currency'] = data['quoteCurrency']['symbol']

    return signal


def update_signal(signal):
    signal['utc_now'] = datetime.utcnow().isoformat()

    query = get_query_latest(signal)
    data = run_query(query)['data']['ethereum']["dexTrades"][0]

    signal['updated_time'] = data['timeInterval']['second']

    signal['updated_price'] = data['quotePrice']

    ret = (signal['updated_price'] - signal['start_price']) / signal['start_price']

    signal['updated_return'] = ret

    return signal



def get_signals_list(filename):
    workbook = openpyxl.load_workbook(filename)

    signals_list = []
    worksheet = workbook.active
    lower_row = get_maxrow(worksheet)
    upper_row = get_maxrow(worksheet) - 10
    for i in range(lower_row, max(1, upper_row), -1):
        signal = get_signal(worksheet, i)
        signals_list.append(signal)

    return signals_list, max(1, upper_row) + 1, lower_row


def update_signals_list(filename):
    initial_signals_list,upper_row,lower_row = get_signals_list(filename)
    signals_list = []
    for signal in initial_signals_list:
        updated_signal = update_signal(signal)
        signals_list.append(updated_signal)
    return signals_list,upper_row,lower_row


def new_signal_detect(filename,x,y):
    signal_list_for_new_signal_detect = []
    workbook = openpyxl.load_workbook(filename)
    worksheet = workbook.active
    new_lower_row = get_maxrow(worksheet)
    old_lower_row = load('rows-new-coin.json')

    if new_lower_row != old_lower_row:
        diff = new_lower_row - old_lower_row
        for i in range(new_lower_row, new_lower_row-diff, -1):
            signal = get_signal(worksheet, i)
            signal_list_for_new_signal_detect.append(signal)
        for i in range(0, diff):
            new_signal_reminder(signal_list_for_new_signal_detect[i],x,y)
    store('rows-new-coin.json', new_lower_row)


def new_signal_reminder(signal,x,y):
    pyautogui.hotkey('win', 'm')
    pyautogui.moveTo(x, y)
    pyautogui.click()
    time.sleep(0.05)
    pyautogui.click()
    time.sleep(0.05)
    pyperclip.copy("【新币： " + signal['base_currency'] + "】" + "\n" +
                   '信号编号： N' + str(signal['row']) + ' (' + str(signal['datetime'])[5:10] + ')' + '\n' +
                   "合约地址： " + signal['token_id'])
    pyautogui.hotkey('ctrl', 'v')
    pyautogui.press('enter')
    time.sleep(0.1)
    pyautogui.hotkey('win', 'm')
    time.sleep(0.1)


def refresh(filename,x,y):
    try:
        new_signals_list, new_upper_row, new_lower_row = update_signals_list(filename)
        data_05 = load("fifty_percent.json")
        data_1 = load("1_time.json")
        data_2 = load("2_times.json")
        data_5 = load("5_times.json")
        data_10 = load("10_times.json")
        for i in range(len(new_signals_list)):
            if new_signals_list[i]["updated_return"] >= 0.5 and  (new_signals_list[i]['row'] not in data_05):
                return_reminder(new_signals_list[i],'50%',x,y)
                # return_reminder(new_signals_list[i], '50%', 280, 342)
                add_data('fifty_percent.json', new_signals_list[i]['row'])
            if new_signals_list[i]["updated_return"] >= 1 and  (new_signals_list[i]['row'] not in data_1):
                return_reminder(new_signals_list[i],'100%',x,y)
                # return_reminder(new_signals_list[i], '100%', 280, 342)
                add_data('1_time.json', new_signals_list[i]['row'])
            if new_signals_list[i]["updated_return"] >= 2 and  (new_signals_list[i]['row'] not in data_2):
                return_reminder(new_signals_list[i],'200%',x,y)
                # return_reminder(new_signals_list[i], '200%', 280, 342)
                add_data('2_times.json', new_signals_list[i]['row'])
            if new_signals_list[i]["updated_return"] >= 5 and  (new_signals_list[i]['row'] not in data_5):
                return_reminder(new_signals_list[i],'500%',x,y)
                # return_reminder(new_signals_list[i], '500%', 280, 342)
                add_data('5_times.json', new_signals_list[i]['row'])
            if new_signals_list[i]["updated_return"] >= 10 and  (new_signals_list[i]['row'] not in data_10):
                return_reminder(new_signals_list[i],'1000%',x,y)
                # return_reminder(new_signals_list[i], '1000%', 280, 342)
                add_data('10_times.json', new_signals_list[i]['row'])

        pyautogui.hotkey('win', 'm')
        pyautogui.moveTo(253, 201)
        pyautogui.click()
        time.sleep(0.05)
        pyautogui.click()
        time.sleep(0.05)
        pyperclip.copy(
            "当前时间： " + str(get_datetime(int(time.time()) * 1000))[5:16] + '\n' +
            '当前追踪： 【新币】第' + str(new_upper_row) + '行 至 第' + str(new_lower_row) + '行 信号'
        )
        pyautogui.hotkey('ctrl', 'v')
        pyautogui.press('enter')
        time.sleep(0.1)
        pyautogui.hotkey('win', 'm')
        time.sleep(0.1)
    except:
        alert(253, 201)

def return_reminder(signal,ret,x,y):
    pyautogui.hotkey('win','m')
    pyautogui.moveTo(x,y)
    pyautogui.click()
    time.sleep(0.05)
    pyautogui.click()
    time.sleep(0.05)
    pyperclip.copy(
        "【新币： " + signal['base_currency'] + "收益超过" + ret + "】" + "\n" +
        '信号编号： N' + str(signal['row']) + ' (' + str(signal['start_time'])[5:10] + ')' + '\n' +
        "当前时间： " + str(get_datetime(int(time.time()) * 1000))[5:16] + '\n' +
        "当前价格： " + str(signal['updated_price']) + '\n' +
        '当前收益： {:.2%}'.format(signal['updated_return']) + '\n'
    )
    pyautogui.hotkey('ctrl', 'v')
    pyautogui.press('enter')
    time.sleep(0.1)
    pyautogui.hotkey('win','m')
    time.sleep(0.1)




def get_query_latest(signal):
    begin_text = "{ethereum(network: " + signal['network'] + """) {dexTrades(options: {limit: 1, desc: "timeInterval.second"}"""
    time_text = "time: {between: [" + "\"" + signal['utc_start'] + "\"" + "," + "\"" + signal['utc_now'] + "\"" + "]}"
    contract_text = "baseCurrency: {is:" + "\"" + signal['token_id'] + "\"" + "}" + \
                    "smartContractAddress: {is:" + "\"" + signal["pair_id"] + "\"" + "}" + ")"
    end_text = """{
                  timeInterval {
                    hour(count: 1)
                    minute(count: 1)
                    second(count: 1)
                  }
                  baseCurrency {
                    symbol
                    address
                  }
                  baseAmount
                  quoteCurrency {
                    symbol
                    address
                  }
                  quoteAmount
                  trades: count
                  quotePrice
                  maximum_price: quotePrice(calculate: maximum)
                  minimum_price: quotePrice(calculate: minimum)
                  open_price: minimum(of: block, get: quote_price)
                  close_price: maximum(of: block, get: quote_price)
                }
              }
            }"""

    query = begin_text + time_text + contract_text + end_text

    return query

def get_query_first(signal):
    begin_text = "{ethereum(network: " + signal['network'] + """) {dexTrades(options: {limit: 1, asc: "timeInterval.second"}"""
    time_text = "time: {between: [" + "\"" + signal['utc_start'] + "\"" + "," + "\"" + signal['utc_now'] + "\"" + "]}"
    contract_text = "baseCurrency: {is:" + "\"" + signal['token_id'] + "\"" + "}" + \
                    "smartContractAddress: {is:" + "\"" + signal["pair_id"] + "\"" + "}" + ")"
    end_text = """{
                  timeInterval {
                    hour(count: 1)
                    minute(count: 1)
                    second(count: 1)
                  }
                  baseCurrency {
                    symbol
                    address
                  }
                  baseAmount
                  quoteCurrency {
                    symbol
                    address
                  }
                  quoteAmount
                  trades: count
                  quotePrice
                  maximum_price: quotePrice(calculate: maximum)
                  minimum_price: quotePrice(calculate: minimum)
                  open_price: minimum(of: block, get: quote_price)
                  close_price: maximum(of: block, get: quote_price)
                }
              }
            }"""

    query = begin_text + time_text + contract_text + end_text

    return query

def run_query(query):  # A simple function to use requests.post to make the API call.
    headers = {'X-API-KEY': 'BQYZxE2l1EL39A7sNIrscIR2noZ03hnU'}
    request = requests.post('https://graphql.bitquery.io/',
                            json={'query': query}, headers=headers)
    if request.status_code == 200:
        return request.json()
    else:
        raise Exception('Query failed and return code is {}.      {}'.format(request.status_code,
                        query))


def alert(x,y):
    pyautogui.hotkey('win','m')
    pyautogui.moveTo(x,y)
    pyautogui.click()
    time.sleep(0.05)
    pyautogui.click()
    time.sleep(0.05)
    pyperclip.copy(
        "当前时间： " + str(get_datetime(int(time.time()) * 1000))[5:16] + '\n' +
        '新币信号追踪运行异常'
    )

    pyautogui.hotkey('ctrl', 'v')
    pyautogui.press('enter')
    time.sleep(0.1)
    pyautogui.hotkey('win','m')
    time.sleep(0.1)



